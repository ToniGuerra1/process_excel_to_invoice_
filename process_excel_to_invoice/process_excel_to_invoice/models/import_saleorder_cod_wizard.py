# -*- coding: utf-8 -*-
import base64
import logging
from io import BytesIO

from odoo import models, fields, api
from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("The openpyxl library was not found in the environment.")


class ImportSaleOrderCODWizard(models.TransientModel):
    _name = 'import.saleorder.cod.wizard'
    _description = 'Wizard to import the payment_reference_to_invoice field into sale.order from Excel.'

    file = fields.Binary(
        string="Excel File",
        required=True,
        help="Upload the XLSX file and specify the column where the order number appears"
    )
    filename = fields.Char(string="Nombre del archivo")
    cod_reference = fields.Char(
        string="Payment Reference",
        required=True,
        help="Text that will be saved in the 'payment_reference_to_invoice' field of the found Sales Orders"
    )
    create_invoice = fields.Boolean(
        string="Create Invoice",
        help="If checked, an invoice will be created for each Sales Order that does not yet have one."
    )
    column_number = fields.Integer(
        string="Column Number",
        default=1,
        required=True,
        help="Column number where the order numbers are located (1 for the first, 2 for the second, etc.)."
    )


    def action_import_cod_reference(self):
        """
        Import COD references from an Excel file in the selected column.
        """
        _logger.info("===== [IMPORT SALE ORDER PAYMENT REFERENCE]... =====")

        if not self.filename or not self.filename.lower().endswith('.xlsx'):
            raise ValidationError("The file must be in .xlsx format.")

        try:
            data = base64.b64decode(self.file or b'')
            workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
        except Exception as e:
            raise ValidationError(f"The XLSX file could not be opened. Error: {e}")

        sheet = workbook.active
        _logger.info(f"XLSX file read successfully. Active sheet: {sheet.title}")

        if self.column_number < 1 or self.column_number > sheet.max_column:
            raise ValidationError(f"The selected column number ({self.column_number}) is out of range. The file has {sheet.max_column} columns.")

        filas_procesadas = 0
        for row_idx in range(2, sheet.max_row + 1):
            # Verificar si la fila está completamente en blanco
            if all(sheet.cell(row=row_idx, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                break

            # Leer el número de pedido desde la columna seleccionada
            saleorder_name = sheet.cell(row=row_idx, column=self.column_number).value
            if not saleorder_name:
                _logger.warning(f"[FILA {row_idx}] There is no order number in the column {self.column_number}. It is skipped.")
                continue

            # Buscar órdenes de venta con ese nombre
            orders = self.env['sale.order'].search([('name', '=', str(saleorder_name))])
            if not orders:
                _logger.warning(f"[FILA {row_idx}] Sale Order '{saleorder_name}' Not found. It is skipped.")
                continue

            # Asignar la referencia COD a las órdenes encontradas
            orders.write({'payment_reference_to_invoice': self.cod_reference})

            # Actualizar payment_reference en las facturas de esas órdenes
            for order in orders:
                if order.invoice_ids:
                    order.invoice_ids.write({'payment_reference': self.cod_reference})
                    _logger.info(f"[FILA {row_idx}] Updated payment_reference in {len(order.invoice_ids)} Invoice(s) for the order {order.name}.")

                # Crear factura si se marcó la opción y la orden no tiene una
                if self.create_invoice and not order.invoice_ids:
                    try:
                        invoices = order._create_invoices()
                        for inv in invoices:
                            inv.action_post()
                        _logger.info(f"[FILA {row_idx}] Invoice(s) created for the order {order.name}.")
                    except Exception as e:
                        _logger.error(f"[FILA {row_idx}] Error creating invoice for the order {order.name}: {e}")

            filas_procesadas += 1
            _logger.info(f"[FILA {row_idx}] It was assigned '{self.cod_reference}' a {len(orders)} sale orders.")

        _logger.info(f"===== [IMPORT SALE ORDER COD] Process completed. Rows processed: {filas_procesadas} =====")

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }